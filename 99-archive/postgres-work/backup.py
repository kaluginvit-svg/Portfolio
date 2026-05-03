"""
Модуль для резервного копирования данных.
Объединяет функциональность сохранения в Excel и SQLite.
Обеспечивает параллельное сохранение данных в локальные резервные копии.
"""

import os
import sqlite3
import logging
from typing import Dict, List, Any, Optional
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill

logger = logging.getLogger(__name__)


class BackupManager:
    """Класс для управления всеми резервными копиями (Excel и SQLite)."""
    
    def __init__(self, excel_path: str = "backup.xlsx", sqlite_path: str = "backup.db"):
        """
        Инициализация менеджера резервного копирования.
        
        Args:
            excel_path: Путь к Excel-файлу (по умолчанию backup.xlsx)
            sqlite_path: Путь к SQLite базе данных (по умолчанию backup.db)
        """
        self.excel_path = excel_path
        self.sqlite_path = sqlite_path
        
        # Excel
        self.workbook = None
        self.worksheet = None
        
        # SQLite
        self.connection = None
    
    # ========== Excel методы ==========
    
    def initialize_excel(self, columns: List[Dict[str, Any]]) -> bool:
        """
        Инициализация Excel-файла с заголовками колонок.
        
        Args:
            columns: Список словарей с информацией о колонках из базы данных
            
        Returns:
            True если файл успешно инициализирован, False в противном случае
        """
        try:
            # Проверяем, существует ли файл
            if os.path.exists(self.excel_path):
                # Загружаем существующий файл
                self.workbook = load_workbook(self.excel_path)
                # Используем первый лист или создаем новый
                if len(self.workbook.sheetnames) > 0:
                    self.worksheet = self.workbook.active
                else:
                    self.worksheet = self.workbook.create_sheet("Data")
            else:
                # Создаем новый файл
                self.workbook = Workbook()
                self.worksheet = self.workbook.active
                self.worksheet.title = "Data"
                
                # Создаем заголовки колонок
                headers = [col['column_name'] for col in columns]
                self.worksheet.append(headers)
                
                # Форматируем заголовки
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF")
                
                for cell in self.worksheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                
                # Автоматически подгоняем ширину колонок
                for idx, col in enumerate(self.worksheet.columns, 1):
                    max_length = 0
                    column_letter = self.worksheet.cell(row=1, column=idx).column_letter
                    for cell in col:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    self.worksheet.column_dimensions[column_letter].width = adjusted_width
                
                # Сохраняем файл
                self.workbook.save(self.excel_path)
                logger.info(f"Создан новый Excel-файл: {self.excel_path}")
            
            return True
            
        except Exception as e:
            logger.error(f"Ошибка инициализации Excel-файла: {e}")
            return False
    
    def save_to_excel(self, all_data: List[Dict[str, Any]], columns: List[Dict[str, Any]]) -> bool:
        """
        Полная перезапись всех данных в Excel-файл.
        Удаляет все старые данные и записывает все данные из базы заново.
        
        Args:
            all_data: Список словарей со всеми данными из таблицы
            columns: Список словарей с информацией о колонках из базы данных
            
        Returns:
            True если данные успешно сохранены, False в противном случае
        """
        try:
            # Инициализируем файл (создаст новый или загрузит существующий)
            if not self.initialize_excel(columns):
                return False
            
            # Загружаем файл заново
            if os.path.exists(self.excel_path):
                self.workbook = load_workbook(self.excel_path)
                self.worksheet = self.workbook.active
            else:
                if not self.initialize_excel(columns):
                    return False
            
            # Получаем порядок колонок из заголовков
            headers = [cell.value for cell in self.worksheet[1]]
            
            # Удаляем все строки данных (кроме заголовка)
            # Удаляем строки снизу вверх, чтобы не сбить индексы
            max_row = self.worksheet.max_row
            if max_row > 1:
                self.worksheet.delete_rows(2, max_row)
            
            # Записываем все данные заново
            for data_row in all_data:
                row_data = []
                for header in headers:
                    # Ищем значение в данных (с учетом регистра)
                    value = None
                    for key, val in data_row.items():
                        if key.lower() == header.lower() or key == header:
                            value = val
                            break
                    
                    # Если значение не найдено, используем пустую строку
                    if value is None:
                        value = ""
                    
                    row_data.append(value)
                
                # Добавляем строку данных
                self.worksheet.append(row_data)
            
            # Сохраняем файл
            self.workbook.save(self.excel_path)
            logger.info(f"Все данные ({len(all_data)} записей) успешно перезаписаны в Excel: {self.excel_path}")
            return True
            
        except Exception as e:
            logger.error(f"Ошибка сохранения данных в Excel: {e}")
            return False
    
    # ========== SQLite методы ==========
    
    def connect_sqlite(self) -> bool:
        """
        Подключение к SQLite базе данных.
        
        Returns:
            True если подключение успешно, False в противном случае
        """
        try:
            self.connection = sqlite3.connect(self.sqlite_path, check_same_thread=False)
            self.connection.row_factory = sqlite3.Row
            logger.info(f"Подключение к SQLite базе данных: {self.sqlite_path}")
            return True
        except sqlite3.Error as e:
            logger.error(f"Ошибка подключения к SQLite: {e}")
            return False
    
    def disconnect_sqlite(self):
        """Закрытие подключения к SQLite базе данных."""
        if self.connection:
            try:
                self.connection.close()
                logger.info("Подключение к SQLite закрыто")
            except sqlite3.Error as e:
                logger.error(f"Ошибка при закрытии SQLite: {e}")
    
    def create_sqlite_table(self, table_name: str, columns: List[Dict[str, Any]]) -> bool:
        """
        Создание таблицы в SQLite на основе структуры таблицы из PostgreSQL.
        
        Args:
            table_name: Имя таблицы
            columns: Список словарей с информацией о колонках из PostgreSQL
            
        Returns:
            True если таблица успешно создана, False в противном случае
        """
        if not self.connection:
            if not self.connect_sqlite():
                return False
        
        try:
            cursor = self.connection.cursor()
            
            # Проверяем, существует ли таблица
            cursor.execute("""
                SELECT name FROM sqlite_master 
                WHERE type='table' AND name=?
            """, (table_name,))
            
            if cursor.fetchone():
                logger.info(f"Таблица {table_name} уже существует в SQLite")
                return True
            
            # Формируем SQL для создания таблицы
            column_definitions = []
            for col in columns:
                col_name = col['column_name']
                col_type = col['data_type']
                is_nullable = col.get('is_nullable', 'YES')
                
                # Преобразуем типы PostgreSQL в типы SQLite
                sqlite_type = self._convert_postgresql_type(col_type)
                
                # Формируем определение колонки
                col_def = f'"{col_name}" {sqlite_type}'
                
                # Добавляем NOT NULL, если колонка обязательна
                if is_nullable == 'NO':
                    col_def += ' NOT NULL'
                
                # Если это ID и есть автоинкремент, добавляем PRIMARY KEY AUTOINCREMENT
                if col_name.lower() == 'id' and col.get('column_default') and 'nextval' in str(col.get('column_default', '')).lower():
                    col_def += ' PRIMARY KEY AUTOINCREMENT'
                elif col_name.lower() == 'id' and not col.get('column_default'):
                    # Если ID без автоинкремента, но обязателен
                    col_def += ' PRIMARY KEY'
                
                column_definitions.append(col_def)
            
            # Создаем таблицу
            create_table_sql = f'''
                CREATE TABLE IF NOT EXISTS "{table_name}" (
                    {', '.join(column_definitions)}
                )
            '''
            
            cursor.execute(create_table_sql)
            self.connection.commit()
            
            logger.info(f"Таблица {table_name} успешно создана в SQLite")
            return True
            
        except sqlite3.Error as e:
            logger.error(f"Ошибка создания таблицы в SQLite: {e}")
            if self.connection:
                self.connection.rollback()
            return False
    
    def _convert_postgresql_type(self, pg_type: str) -> str:
        """
        Преобразование типа данных PostgreSQL в тип SQLite.
        
        Args:
            pg_type: Тип данных PostgreSQL
            
        Returns:
            Тип данных SQLite
        """
        type_mapping = {
            'integer': 'INTEGER',
            'int4': 'INTEGER',
            'bigint': 'INTEGER',
            'int8': 'INTEGER',
            'smallint': 'INTEGER',
            'serial': 'INTEGER',
            'bigserial': 'INTEGER',
            'real': 'REAL',
            'double precision': 'REAL',
            'float': 'REAL',
            'numeric': 'REAL',
            'decimal': 'REAL',
            'money': 'REAL',
            'boolean': 'INTEGER',
            'bool': 'INTEGER',
            'date': 'TEXT',
            'time': 'TEXT',
            'timestamp': 'TEXT',
            'timestamp without time zone': 'TEXT',
            'timestamp with time zone': 'TEXT',
            'text': 'TEXT',
            'varchar': 'TEXT',
            'character varying': 'TEXT',
            'char': 'TEXT',
            'character': 'TEXT',
            'json': 'TEXT',
            'jsonb': 'TEXT',
            'uuid': 'TEXT',
            'bytea': 'BLOB'
        }
        
        pg_type_lower = pg_type.lower()
        
        if pg_type_lower in type_mapping:
            return type_mapping[pg_type_lower]
        
        for pg_key, sqlite_type in type_mapping.items():
            if pg_key in pg_type_lower:
                return sqlite_type
        
        return 'TEXT'
    
    def save_to_sqlite(self, table_name: str, data: Dict[str, Any]) -> bool:
        """
        Вставка данных в таблицу SQLite.
        
        Args:
            table_name: Имя таблицы
            data: Словарь с данными для вставки
            
        Returns:
            True если данные успешно вставлены, False в противном случае
        """
        if not self.connection:
            if not self.connect_sqlite():
                return False
        
        if not data:
            logger.error("Нет данных для вставки в SQLite")
            return False
        
        try:
            cursor = self.connection.cursor()
            
            # Формируем SQL запрос
            columns = list(data.keys())
            placeholders = ', '.join(['?' for _ in columns])
            column_names = ', '.join([f'"{col}"' for col in columns])
            values = tuple(data.values())
            
            insert_sql = f'INSERT INTO "{table_name}" ({column_names}) VALUES ({placeholders})'
            
            cursor.execute(insert_sql, values)
            self.connection.commit()
            
            logger.info(f"Данные успешно сохранены в SQLite: {table_name}")
            return True
            
        except sqlite3.Error as e:
            logger.error(f"Ошибка вставки данных в SQLite: {e}")
            if self.connection:
                self.connection.rollback()
            return False
    
    # ========== Универсальные методы ==========
    
    def initialize(self, table_name: str, columns: List[Dict[str, Any]]) -> bool:
        """
        Инициализация всех резервных копий.
        
        Args:
            table_name: Имя таблицы
            columns: Список словарей с информацией о колонках из базы данных
            
        Returns:
            True если все резервные копии успешно инициализированы, False в противном случае
        """
        excel_ok = self.initialize_excel(columns)
        sqlite_ok = False
        
        if self.connect_sqlite():
            sqlite_ok = self.create_sqlite_table(table_name, columns)
        
        return excel_ok and sqlite_ok
    
    def save_data(self, table_name: str, data: Dict[str, Any], columns: List[Dict[str, Any]], db_connection=None) -> bool:
        """
        Автоматическое сохранение данных во все резервные копии (Excel и SQLite).
        Выполняется параллельно и не блокирует основную работу.
        Для Excel получает все данные из базы и полностью перезаписывает файл.
        
        Args:
            table_name: Имя таблицы
            data: Словарь с данными для сохранения (используется только для SQLite)
            columns: Список словарей с информацией о колонках из базы данных
            db_connection: Объект Database для получения всех данных из таблицы (обязателен для Excel)
            
        Returns:
            True если данные успешно сохранены хотя бы в одну копию, False в противном случае
        """
        # Убеждаемся, что таблица SQLite существует
        if not self.connection:
            self.connect_sqlite()
        
        if self.connection:
            cursor = self.connection.cursor()
            cursor.execute("""
                SELECT name FROM sqlite_master 
                WHERE type='table' AND name=?
            """, (table_name,))
            if not cursor.fetchone():
                self.create_sqlite_table(table_name, columns)
        
        # Сохраняем данные параллельно в обе резервные копии
        excel_ok = False
        sqlite_ok = False
        
        # Сохранение в Excel - получаем все данные из базы и перезаписываем файл
        try:
            if db_connection:
                # Получаем все данные из таблицы
                # Пытаемся отсортировать по первой колонке (обычно это ID)
                first_column = columns[0]["column_name"]
                try:
                    query = f'SELECT * FROM "{table_name}" ORDER BY "{first_column}"'
                    all_data = db_connection.execute_query(query)
                except:
                    # Если сортировка не удалась, получаем данные без сортировки
                    query = f'SELECT * FROM "{table_name}"'
                    all_data = db_connection.execute_query(query)
                
                excel_ok = self.save_to_excel(all_data, columns)
                logger.info(f"Excel файл обновлен: загружено {len(all_data)} записей из таблицы {table_name}")
            else:
                logger.warning("Объект Database не передан, Excel резервная копия не обновлена")
        except Exception as e:
            logger.error(f"Ошибка сохранения в Excel: {e}")
        
        # Сохранение в SQLite (добавляем только новую запись)
        try:
            sqlite_ok = self.save_to_sqlite(table_name, data)
        except Exception as e:
            logger.error(f"Ошибка сохранения в SQLite: {e}")
        
        # Возвращаем True, если хотя бы одна копия успешна
        return excel_ok or sqlite_ok
    
    def close(self):
        """Закрытие всех подключений и файлов."""
        # Закрываем Excel
        if self.workbook:
            try:
                self.workbook.close()
            except:
                pass
        
        # Закрываем SQLite
        self.disconnect_sqlite()
